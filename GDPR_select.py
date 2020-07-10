import pandas as pd
from openpyxl import load_workbook
import requests
import time
import os
import configparser

proDir = os.path.split(os.path.realpath(__file__))[0]
configPath = os.path.join(proDir, "config.ini")

class GDPR_select():
    def __init__(self): #初始化白名单
        pass

    def filte_repeat(self, path, raw_sheetname, result_sheetname): #过滤重复数据
        '''
        :param path: xlsx文件路径
        :param raw_sheetname: 被筛选数据原sheet表名称
        :param result_sheetname: 筛选数据后存储的sheet表名称
        :return: 
        '''
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))  # 读取Excel中Sheet1中的数据
        # pd.DataFrame().to_excel(raw_sheetname)
        # print(data) # 查看读取数据内容

        re_row = data.duplicated()  # 查看是否有重复行
        print(re_row)

        no_re_row = data.drop_duplicates()  # 查看去除重复行的数据
        print(no_re_row)

        write = pd.ExcelWriter(path, engine='openpyxl')
        write.book = load_workbook(path)  # 保留原sheet页的数据

        no_re_row.to_excel(write, index=None, sheet_name=result_sheetname)
        write.close()

    def filte_colm(path, raw_sheetname, colmname):   #过滤某列数据
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))
        no_re_by_colm = data.drop_duplicates(subset=colmname)

    def select_colm(self, path, raw_sheetname, colmnum):
        # list_white = []
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))
        print(data)
        # colmvalue = np.array()
        col = data.iloc[:, colmnum].values
        print(col)
        print(type(col))
        return col


    def select(self, path, raw_sheetname, colname, condition, result_sheetname):
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))

        select_result = data[data[colname] == condition]
        print(select_result)

        write = pd.ExcelWriter(path, engine='openpyxl')
        write.book = load_workbook(path)  # 保留原sheet页的数据

        select_result.to_excel(write, index=None, sheet_name=result_sheetname)
        write.close()

    def select_by_list(self, path, raw_sheetname, column_name, list1, result_sheetname):
        '''
        :param path: 文件路径
        :param raw_sheetname: 数据来源的sheet表
        :param column_name: 列值
        :param list1: 白名单列表或者欧盟列表等（用于核对的列表）
        :param result_sheetname:
        :return:
        '''
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))
        result_data = data.loc[~data[column_name].isin(list1)]
        write = pd.ExcelWriter(path, engine='openpyxl')
        write.book = load_workbook(path)  # 保留原sheet页的数据

        result_data.to_excel(write, index=None, sheet_name=result_sheetname)
        write.close()


    def select_com(self, path, raw_sheetname, colmnum, result_sheetname):
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))
        print(data)
        # list2 = GDPR_select.select_colm('过滤重复行.xlsx', '筛选重复数据结果', 0)
        list2 = GDPR_select.select_colm(self, path, raw_sheetname, colmnum)
        print(list2)
        write = pd.ExcelWriter(path, engine='openpyxl')
        write.book = load_workbook(path)
        list3 = []
        for i in list2:
            com = str(i).split(':')[0]
            list3.append(com)
            print(list3)
            data1 = pd.DataFrame({"网址": list3})
            data['网址'] = data1
        print(data)
        data.to_excel(write, sheet_name=result_sheetname, index=None)
        write.save()

    def query(ip, note = ''):
        url = "http://ip-api.com/json/%s" % ip
        response = requests.get(url)
        print(response.text)
        rep = {}
        # if response.status_code is 200:
        #     return 1
        if not response.text:
            print("This is null, next one ...")
            return
        else:
            rep = response.json()

        print("*************************")
        print("您查询的IP地址 %s 来源地是：" % (rep.get('query')))
        print("国家：%s" % (rep.get('country')))
        print("城市：%s" % (rep.get('city')))
        print("经纬度坐标：%s,%s" % (rep.get('lat'), rep.get('lon')))
        print("运营商编号：%s" % (rep.get('as')))
        print("ISP服务商：%s" % (rep.get('isp')))
        print("*************************")
        return "%s 请确认：%s - %s" % (note, rep.get('country'), rep.get('isp'))
        
    def IP_check(self, path, raw_sheetname, list_IP, result_sheetname):
        data = pd.DataFrame(pd.read_excel(path, raw_sheetname))
        write = pd.ExcelWriter(path, engine='openpyxl')
        write.book = load_workbook(path)
        T_type = select.select_colm(path, raw_sheetname, 3)
        print(T_type)
        i = 0
        result = []
        for p in list_IP:
            i += 1
            print(p, i)
            if T_type[i-1]=='HTTP':
                add = GDPR_select.query(p, 'HTTP')
            else:
                add = GDPR_select.query(p)
            if add.find('None') > 0:
                add = ''

            time.sleep(2)
            result.append(add)
        print(result)
        data1 = pd.DataFrame({'备注':result})
        data['备注'] = data1

        # write_add_col(path, result, '查询结果')
        data.to_excel(write, sheet_name=result_sheetname, index=None)
        write.save()

    def main_EU(self, path, white_path, EU_NAME):
        GDPR_select.filte_repeat(self, path, '全部数据', '筛选重复数据结果')

        # GDPR_select.select(path, '筛选重复数据结果', 'Country', 'Local Address', '筛选去掉本地数据')
        GDPR_select.select_by_list(self, path, '筛选重复数据结果', 'Country', ['Local Address'], '筛选去掉本地数据')

        list1 = GDPR_select.select_colm(self, white_path, 'GDPR测试应用白名单', 1)
        GDPR_select.select_com(self, path, '筛选去掉本地数据', 0, '已过滤用于核对白名单数据')
        GDPR_select.select_by_list(self, path, '已过滤用于核对白名单数据', '网址', list1, '已过滤白名单')

        GDPR_select.select_by_list(self, path, '已过滤白名单', 'Country', EU_NAME, '非欧盟国家过滤')
        # 查询IP地址后输出
        UNEU_IP_list = GDPR_select.select_colm(self, path, '非欧盟国家过滤', 1)
        print(UNEU_IP_list)
        GDPR_select.IP_check(self, path, '非欧盟国家过滤', UNEU_IP_list, '非欧盟国家跨境传输数据')

        select.select(path, '已过滤白名单', 'HTTP/HTTPS', 'HTTP', 'HTTP传输')
        # 查询IP地址后输出
        IP_lsit = GDPR_select.select_colm(self, path, 'HTTP传输', 1)
        print(IP_lsit)
        GDPR_select.IP_check(self, path, 'HTTP传输', IP_lsit, 'HTTP传输数据')

    def main_AU(self):
        #筛选去掉重复数据
        GDPR_select.filte_repeat(self, path, '全部数据', '筛选重复数据结果')
        #筛选去掉本地数据
        GDPR_select.select_by_list(self, path, '筛选重复数据结果', 'Country', ['Local Address'], '筛选去掉本地数据')
        #筛选白名单
        list1 = GDPR_select.select_colm(self, white_path, 'GDPR测试应用白名单', 1)
        GDPR_select.select_com(self, path, '筛选去掉本地数据', 0, '已过滤用于核对白名单数据')
        GDPR_select.select_by_list(self, path, '已过滤用于核对白名单数据', '网址', list1, '已过滤白名单')

        GDPR_select.select_by_list(self, path, '已过滤白名单', 'HTTP/HTTPS', AU_Security, '过滤安全数据')
        IP_lsit = GDPR_select.select_colm(self, path, '过滤安全数据', 1)
        print(IP_lsit)
        GDPR_select.IP_check(self, path, '过滤安全数据', IP_lsit, '非安全传输数据')

    def main_flow(self, Flag):  #主函数
        try:
            if Flag.lower() == 'EU'.lower():
                select.main_EU(path, white_path, EU_NAME)
            elif Flag.lower() == 'AU'.lower():
                select.main_AU()
            else:
                print("输入错误，请重新输入")
                Flag = str(input("请输入 EU 或者 AU\n"))
                GDPR_select.main_flow(self,Flag)

        except ValueError as e:
            print("输入错误，请重新输入")


class ReadConfig:
    def __init__(self):
        self.cf = configparser.ConfigParser()
        self.cf.read(configPath,encoding='UTF-8')
        # print(cf)

    def get_path(self, param):
        value = self.cf.get("PATH",param)
        print(value)
        return value

if __name__=="__main__":
    #获取配置文件钟报告路径
    read_ini = ReadConfig()
    path = read_ini.get_path('path')
    white_path = read_ini.get_path('white_path')
    
    # #报告路径参数
    # path = r'D:\Auto-test\GDPR_report_select_V1.0\CCPA-V047版本_US_report.xlsx'
    # white_path = r'D:\Auto-test\GDPR_report_select_V1.0\海外TV产品GDPR测试应用白名单V1.0(1).xlsx'
    Flag = str(input("请输入 EU 或者 AU\n"))
    
    #欧盟列表
    EU_NAME = ['Austria', 'Belgium', 'Bulgaria', 'Cyprus', 'Croatia', 'Czechia', 'Denmark', 'Estonia', 'Finland',
               'France', 'Germany',
               'Greece', 'Hungary', 'Ireland', 'Italy', 'Latvia', 'Lithuania', 'Luxembourg', 'Malta', 'Netherlands',
               'Poland', 'Portugal',
               'Romania', 'Slovakia', 'Slovenia', 'Spain', 'Sweden']  # 'United Kingdom'] 应该已脱离欧盟，将英国过滤去掉
    AU_Security = ['SSL', 'TLS1.2', 'TLS1.0']

    # EU = {'Austria': '奥地利', 'Belgium': '比利时', 'Bulgaria': '保加利亚', 'Cyprus': '塞浦路斯', 'Croatia': '克罗地亚', 'Czechia': '捷克',
          # 'Denmark': '丹麦', 'Estonia': '爱沙尼亚', 'Finland': '芬兰', 'France': '法国', 'Germany': '德国', 'Greece': '希腊',
    #       'Hungary': '匈牙利',
    #       'Ireland': '爱尔兰', 'Italy': '意大利', 'Latvia': '拉脱维亚', 'Lithuania': '立陶宛', 'Luxembourg': '卢森堡', 'Malta': '马耳他',
    #       'Netherlands': '荷兰', 'Poland': '波兰', 'Portugal': '葡萄牙', 'Romania': '罗马尼亚', 'Slovakia': '斯洛伐克',
    #       'Slovenia': '斯洛文尼亚',
    #       'Spain': '西班牙', 'Sweden': '瑞典', 'United Kingdom': '英国'}

    select = GDPR_select()
    select.main_flow(Flag)



